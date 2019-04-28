# coding=utf-8
from openpyxl import load_workbook
import numpy as np
import pandas as pd
from tqdm import tqdm
import time
import os

DATA_PATH = './data/data2.xlsx'
SHEET1_NAME = 'biao1'
SHEET2_NAME = 'biao2'
R_ROW_INDEX = 5  # 调幅参数r在excel表中的行索引
EN_ROW_INDEX = 3  # EN载荷参数在excel表中的行索引
EN_START_COLUMN_INDEX = 2  # EN载荷参数列起始索引
EN_END_COLUMN_INDEX = 13  # EN载荷参数列结束索引
MATRIX_A_COUNT = 4  # 矩阵Ai的个数
MATRIX_A_ROW_NUM = 66  # 矩阵Ai行数
MATRIX_A_COLUMN_NUM = 11  # 矩阵Ai列数
MATRIX_A_START_COLUMN_INDEX_LIST = [2, 15, 28, 41]  # 矩阵Ai在excel表中开始的列索引列表
MATRIX_A_START_ROW_INDEX = 7  # 矩阵Ai在excel表中开始的行索引
MATRIX_A_END_ROW_INDEX = MATRIX_A_START_ROW_INDEX + MATRIX_A_ROW_NUM  # 矩阵Ai在excel表中结束的行索引
DELTA_COL_INDEX = 55  # delta在excel表中的列索引
DELTA_ROW_START_INDEX = 3  # delta在excel表中的行起始索引
DELTA_ROW_END_INDEX = MATRIX_A_ROW_NUM + DELTA_ROW_START_INDEX  # delta在excel表中的行终止索引
R_MIN_VALUE = 0.5  # 参数最大值
R_MAX_VALUE = 1.5  # 参数最小值
R_OPTIMISE_STEP = 0.5  # 参数优化步长


# 获取数据
def getData():
    # 获取数据
    data = load_workbook(DATA_PATH)

    # 获取sheet1
    table1 = data[SHEET1_NAME]

    # 获取EN载荷
    F = np.zeros(MATRIX_A_COLUMN_NUM)
    for col in range(EN_START_COLUMN_INDEX, EN_END_COLUMN_INDEX):
        F[col - EN_START_COLUMN_INDEX] = table1.cell(row=EN_ROW_INDEX, column=col).value
    # print(F, F.shape)

    # 获取矩阵A1-A4及R
    A = np.zeros((MATRIX_A_COUNT, MATRIX_A_ROW_NUM, MATRIX_A_COLUMN_NUM))
    R = np.ones((MATRIX_A_COUNT, MATRIX_A_COLUMN_NUM, 1))

    for i in range(MATRIX_A_COUNT):
        col = 0
        for j in range(MATRIX_A_START_COLUMN_INDEX_LIST[i], MATRIX_A_START_COLUMN_INDEX_LIST[i] + MATRIX_A_COLUMN_NUM):
            A[i][:, col] = np.array([table1.cell(row=row, column=j).value for row in
                                     range(MATRIX_A_START_ROW_INDEX, MATRIX_A_END_ROW_INDEX)])
            R[i, col, 0] = table1.cell(row=R_ROW_INDEX, column=j).value
            col += 1
    # print(A.shape, A[0][:, 1], R)

    # 获取sheet2
    table2 = data[SHEET2_NAME]

    # 读取实际值
    D = np.zeros(MATRIX_A_ROW_NUM)
    for row in range(DELTA_ROW_START_INDEX, DELTA_ROW_END_INDEX):
        D[row - DELTA_ROW_START_INDEX] = table2.cell(row=row, column=DELTA_COL_INDEX).value
    # print(D, D.shape)
    return F, A, D, R


# 保存数据
def saveData(R, path, loss, num):
    # 获取源数据
    data = load_workbook(DATA_PATH)
    # 获取sheet1
    table1 = data[SHEET1_NAME]
    for i in range(MATRIX_A_COUNT):
        col = 0
        for j in range(MATRIX_A_START_COLUMN_INDEX_LIST[i], MATRIX_A_START_COLUMN_INDEX_LIST[i] + MATRIX_A_COLUMN_NUM):
            table1.cell(row=R_ROW_INDEX, column=j).value = R[i, col, 0]
            col += 1
    if not os.path.exists(path):
        os.mkdir(path)
    data.save(os.path.join(path, 'result_%.2f_%d_%s.xlsx' % (
        loss, num, time.strftime("%Y%m%d%H%M", time.localtime(time.time())))))


# 计算比值
def calRate(F, A, D, R):
    X = np.zeros((MATRIX_A_ROW_NUM, 1))
    for i in range(MATRIX_A_COUNT):
        m = np.mat(F * A[i] * 0.21) * np.mat(R[i])
        X += 50 * np.power(np.power(m, 2), 1.75)
    X = np.power(X / 200, 1 / 3.5)[:, 0]
    loss = np.log(np.sum(np.square(np.abs(X - D))))
    rate = X / D
    return loss, rate


# 计算比值
def calRateWithInterSum(F, A, D, R):
    X = np.zeros((MATRIX_A_ROW_NUM, 1))
    interSums = np.zeros((MATRIX_A_ROW_NUM, MATRIX_A_COUNT))
    for i in range(MATRIX_A_COUNT):
        m = np.mat(F * A[i] * 0.21) * np.mat(R[i])
        X += 50 * np.power(np.power(m, 2), 1.75)
        interSums[:, i] = np.sum(F * A[i] * 0.21, axis=1)
    X = np.power(X / 200, 1 / 3.5)[:, 0]
    loss = np.log(np.sum(np.square(np.abs(X - D))))
    rate = X / D
    return interSums, loss, rate


# 优化参数-
def explore(F, A, D, _R, minValue=R_MIN_VALUE, maxValue=R_MAX_VALUE, step=R_OPTIMISE_STEP):
    R = _R.copy()
    # 计算初始loss和比值
    interSums, loss, rate = calRateWithInterSum(F, A, D, R)
    # 找出rate<1的行
    rate = pd.Series(rate)
    optList = rate[rate < 1].index.tolist()
    optCount = len(optList)
    # 优化结束标志
    overFlag = False
    # 遍历
    for row in optList:
        interSum = pd.Series(abs(interSums[row, :]))
        absMinIndex = interSum.sort_values().index.tolist()
        # 绝对值最小值索引i，对应Ai
        # absMinIndex = interSum[interSum == min(interSum)].index.tolist()
        for i in absMinIndex:
            for col in range(MATRIX_A_COLUMN_NUM):
                if A[i, row, col] > 0:
                    op = 1
                else:
                    op = -1
                r = r_origin = r_current = R[i, col, 0]

                while r_current > minValue and r_current < maxValue:
                    r_current += op * step
                    R_tmp = R.copy()
                    R_tmp[i, col, 0] = r_current
                    _, loss_current, rate_current = calRateWithInterSum(F, A, D, R_tmp)
                    count = len(rate_current[rate_current < 1])
                    if count > optCount:
                        pass
                    elif count == optCount:
                        if loss_current > loss:
                            pass
                        else:
                            r_origin = r_current
                            break
                    else:
                        r_origin = r_current
                        break
                if r == r_origin:
                    continue
                R[i, col, 0] = r_origin
                _, loss_current, rate_current = calRateWithInterSum(F, A, D, R)
                optCount = len(rate_current[rate_current < 1])
                if optCount == 0:
                    overFlag = True
                    break
            if overFlag:
                break
        if overFlag:
            break
    return R


# 优化
def optimise(R):
    step = R_OPTIMISE_STEP
    minValue = R_MIN_VALUE
    maxValue = R_MAX_VALUE
    roundThreshold = 4  # rate不再变化的迭代次数阈值，大于该阈值则增加上界
    maxValueStep = 0.5  # 上界增加步长
    loss_last, rate_last = calRate(F, A, D, R)
    count_last = len(rate_last[rate_last < 1])
    # 记录rate不再变化的round次数
    roundCounter = 0
    for _ in tqdm(range(100)):
        R = explore(F, A, D, R, minValue, maxValue, step)
        loss_current, rate_current = calRate(F, A, D, R)
        count_current = len(rate_current[rate_current < 1])
        if count_current == 0:
            print('early stop')
            break
        if count_current == count_last:
            roundCounter += 1
        else:
            count_last = count_current
        # rate不再变化的迭代次数阈值，大于该阈值则增加上界
        if roundCounter == roundThreshold:
            maxValue += maxValueStep
            roundCounter = 0
    return R


# 测试
def test():
    # 获取数据：已知矩阵F，已知矩阵Ai和参数R
    F, A, D, R = getData()
    # print(F.shape, A.shape, D.shape, R.shape)

    loss, rate = calRate(F, A, D, R)
    print(loss, rate, len(rate[rate < 1]))

    R = np.ones((MATRIX_A_COUNT, MATRIX_A_COLUMN_NUM, 1)) * 0.5
    loss, rate = calRate(F, A, D, R)
    print(loss, rate, len(rate[rate < 1]))

    # 保存数据
    saveData(R, './result2', loss, len(rate[rate < 1]))


if __name__ == "__main__":
    # 获取数据：已知矩阵F，已知矩阵Ai和参数R
    F, A, D, R = getData()
    print(F.shape, A.shape, D.shape, R.shape)

    # 寻找最优参数
    R = optimise(R)

    # 保存最终的结果
    loss, rate = calRate(F, A, D, R)
    saveData(R, './result2', loss, len(rate[rate < 1]))

    print(loss, len(rate[rate < 1]))
