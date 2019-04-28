# coding=utf-8
from openpyxl import load_workbook
import numpy as np
from tqdm import tqdm
import time
import os
import logging

from Fitness import Fitness
from GA import GA

DATA_PATH = './data/data.xlsx'
SHEET1_NAME = 'biao1'
SHEET2_NAME = 'biao2'
R_ROW_INDEX = 5  # 调幅参数r在excel表中的行索引
EN_ROW_INDEX = 3  # EN载荷参数在excel表中的行索引
EN_START_COLUMN_INDEX = 2  # EN载荷参数列起始索引
EN_END_COLUMN_INDEX = 14  # EN载荷参数列结束索引
MATRIX_A_COUNT = 4  # 矩阵Ai的个数
MATRIX_A_ROW_NUM = 66  # 矩阵Ai行数
MATRIX_A_COLUMN_NUM = 12  # 矩阵Ai列数
MATRIX_A_START_COLUMN_INDEX_LIST = [2, 16, 30, 44]  # 矩阵Ai在excel表中开始的列索引列表
MATRIX_A_START_ROW_INDEX = 7  # 矩阵Ai在excel表中开始的行索引
MATRIX_A_END_ROW_INDEX = MATRIX_A_START_ROW_INDEX + MATRIX_A_ROW_NUM  # 矩阵Ai在excel表中结束的行索引
DELTA_COL_INDEX = 59  # delta在excel表中的列索引
DELTA_ROW_START_INDEX = 3  # delta在excel表中的行起始索引
DELTA_ROW_END_INDEX = MATRIX_A_ROW_NUM + DELTA_ROW_START_INDEX  # delta在excel表中的行终止索引


# 获取数据
def getData():
    # 获取数据
    data = load_workbook(DATA_PATH)

    # 获取sheet1
    table1 = data[SHEET1_NAME]

    # 获取EN载荷
    F = np.zeros(MATRIX_A_COLUMN_NUM)
    for col in range(EN_START_COLUMN_INDEX, EN_END_COLUMN_INDEX):
        F[col - EN_START_COLUMN_INDEX] = table1.cell(row=EN_ROW_INDEX, column=col).value
    # print(F, F.shape)

    # 获取矩阵A1-A4及R
    A = np.zeros((MATRIX_A_COUNT, MATRIX_A_ROW_NUM, MATRIX_A_COLUMN_NUM))
    R = np.ones((MATRIX_A_COUNT, MATRIX_A_COLUMN_NUM, 1))

    for i in range(MATRIX_A_COUNT):
        col = 0
        for j in range(MATRIX_A_START_COLUMN_INDEX_LIST[i], MATRIX_A_START_COLUMN_INDEX_LIST[i] + MATRIX_A_COLUMN_NUM):
            A[i][:, col] = np.array([table1.cell(row=row, column=j).value for row in
                                     range(MATRIX_A_START_ROW_INDEX, MATRIX_A_END_ROW_INDEX)])
            R[i, col, 0] = table1.cell(row=R_ROW_INDEX, column=j).value
            col += 1
    # print(A.shape, A[0][:, 1], R)

    # 获取sheet2
    table2 = data[SHEET2_NAME]

    # 读取实际值
    D = np.zeros(MATRIX_A_ROW_NUM)
    for row in range(DELTA_ROW_START_INDEX, DELTA_ROW_END_INDEX):
        D[row - DELTA_ROW_START_INDEX] = table2.cell(row=row, column=DELTA_COL_INDEX).value
    # print(D, D.shape)
    return F, A, D, R


# 保存数据
def saveData(R, path, loss, num):
    # 获取源数据
    data = load_workbook(DATA_PATH)
    # 获取sheet1
    table1 = data[SHEET1_NAME]
    for i in range(MATRIX_A_COUNT):
        col = 0
        for j in range(MATRIX_A_START_COLUMN_INDEX_LIST[i], MATRIX_A_START_COLUMN_INDEX_LIST[i] + MATRIX_A_COLUMN_NUM):
            table1.cell(row=R_ROW_INDEX, column=j).value = R[i, col, 0]
            col += 1
    if not os.path.exists(path):
        os.mkdir(path)
    data.save(os.path.join(path, 'result_%.2f_%d_%s.xlsx' % (
        loss, num, time.strftime("%Y%m%d%H%M", time.localtime(time.time())))))


# 计算比值
def calRate(F, A, D, R):
    X = np.zeros((MATRIX_A_ROW_NUM, 1))
    for i in range(MATRIX_A_COUNT):
        m = np.mat(F * A[i] * 0.21) * np.mat(R[i])
        X += 50 * np.power(np.power(m, 2), 1.75)
    X = np.power(X / 200, 1 / 3.5)[:, 0]
    loss = np.log(np.sum(np.square(np.abs(X - D))))
    rate = X / D
    return loss, rate

# 计算比值
def calRateWithInterSum(F, A, D, R):
    X = np.zeros((MATRIX_A_ROW_NUM, 1))
    interSum = np.zeros((MATRIX_A_ROW_NUM, MATRIX_A_COUNT))
    for i in range(MATRIX_A_COUNT):
        m = np.mat(F * A[i] * 0.21) * np.mat(R[i])
        X += 50 * np.power(np.power(m, 2), 1.75)
        interSum[:, i] = np.sum(F * A[i] * 0.21, axis=1)
        # print(np.sum(F * A[i] * 0.21, axis=1))
    X = np.power(X / 200, 1 / 3.5)[:, 0]
    loss = np.log(np.sum(np.square(np.abs(X - D))))
    rate = X / D
    return loss, rate


# 自定义适应度函数
class CustomFitness(Fitness):
    def __init__(self, F, A, D):
        self.F = F
        self.A = A
        self.D = D

    # 适应度计算
    def fitnessFunc(self, decodedGene):
        R = decodedGene.reshape((MATRIX_A_COUNT, MATRIX_A_COLUMN_NUM, 1))
        loss, rate = calRate(F, A, D, R)
        return 100 / (loss + len(rate[rate < 1]))


def test():
    # 获取数据：已知矩阵F，已知矩阵Ai和参数R
    F, A, D, R = getData()
    # print(F.shape, A.shape, D.shape, R.shape)

    loss, rate = calRate(F, A, D, R)
    print(loss, rate)

    R = np.ones((MATRIX_A_COUNT, MATRIX_A_COLUMN_NUM, 1)) * 0.5
    loss, rate = calRate(F, A, D, R)
    print(loss, rate)

    # 保存数据
    saveData(R, './result1', loss, len(rate[rate < 1]))


if __name__ == "__main__":
    # 获取数据：已知矩阵F，已知矩阵Ai和参数R
    F, A, D, R = getData()
    # print(F.shape, A.shape, D.shape, R.shape)

    # 适应函数度量
    fitnessClass = CustomFitness(F, A, D)

    # 使用遗传算法寻找最优解
    ga = GA(crossoverRate=0.9, mutationRate=0.2, populationSize=200, boundaryList=[[0.5, 1.5]] * 48, delta=0.05,
            fitnessClass=fitnessClass, fitnessThreshold=6, similarityThreshold=30, CF=2, punish=0.01,
            showBestFitness=True, showTotalFitness=False, earlyStopRoundThreshold=100)

    logging.info(ga.parameters())
    logging.info('#generation_%d best fitness: %f' % (ga._generation, ga._bestLife._fitness))
    for n in tqdm(range(500)):
        # 生成新一代
        ga.getNewGeneration()
        logging.info('#generation_%d best fitness: %f' % (ga._generation, ga._bestLife._fitness))
        if ga._earlyStop:
            logging.info('early stop')
            break
    logging.info('crossover count: %d' % ga._crossoverCount)
    logging.info('mutation count: %d' % ga._mutationCount)

    decodedGene = ga.decodedOneGene(ga._bestLife._gene)
    logging.info('encode gene = \n %s \n' % ga._bestLife._gene)
    logging.info('decode gene = \n %s \n' % decodedGene)
    logging.info('fitness = %s \n' % ga._bestLife._fitness)

    # 可视化历史最佳适应度
    ga.plotHistory()
    ga._showTotalFitness = True
    ga.plotHistory()

    # 得到寻找到的最优解
    R = decodedGene.reshape((MATRIX_A_COUNT, MATRIX_A_COLUMN_NUM, 1))
    loss, rate = calRate(F, A, D, R)
    logging.info('LOSS = %f, len(rate > 1) = %d' % (loss, len(rate[rate < 1])))

    # 将参数写入excel表中
    saveData(R, './result1', loss, len(rate[rate < 1]))
